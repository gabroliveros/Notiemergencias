# -*- coding: utf-8 -*-
"""
Genera los dos XLSForm que se importan manualmente en KoboToolbox
(Proyectos -> New -> Import an XLSForm):

  1) noticias_emergencia_vzla.xlsx    -> formulario de noticias relevantes
  2) metricas_alerta_vzla.xlsx        -> formulario de métricas/nivel de alerta

Los choices de 'estado' se generan desde diccionarios.ESTADOS para que
coincidan exactamente con lo que usa el resto del pipeline. Los choices de
'tipo_evento' coinciden con MAPA_EVENTO_CANONICO de metricas_alerta.py
(hoy: sismo, inundacion, deslave, lluvias — sequía y huracán quedan
pendientes hasta que se agreguen sus palabras clave en diccionarios.py).

Uso:
    python -m kobo.xlsform_builder
Genera ambos archivos en la carpeta actual.
"""

import os
import unicodedata

import openpyxl

from procesamiento.diccionarios import ESTADOS
from procesamiento.metricas_alerta import MAPA_EVENTO_CANONICO

TIPOS_EVENTO = sorted(set(MAPA_EVENTO_CANONICO.values()))
NIVELES_ALERTA = ['rojo', 'naranja', 'amarillo', 'verde']


def nombre_choice(texto):
    """Convierte un label a un 'name' de choice válido para XLSForm (sin
    espacios ni tildes): 'la guaira' -> 'la_guaira'. Pública porque también
    la usa gestionar_formularios/subir_a_kobo.py al preparar submissions."""
    sin_tildes = unicodedata.normalize('NFKD', texto)
    sin_tildes = ''.join(c for c in sin_tildes if not unicodedata.combining(c))
    return sin_tildes.lower().strip().replace(' ', '_')


def _escribir_hoja(wb, nombre_hoja, encabezados, filas):
    ws = wb.create_sheet(nombre_hoja)
    ws.append(encabezados)
    for fila in filas:
        ws.append(fila)
    return ws


def _hoja_choices_comun(wb):
    """Genera la hoja 'choices' compartida por ambos formularios (estados,
    tipos de evento, niveles de alerta, sí/no)."""
    filas = []
    for estado in ESTADOS:
        filas.append(['estados', nombre_choice(estado), estado.capitalize()])
    for evento in TIPOS_EVENTO:
        filas.append(['tipos_evento', evento, evento.capitalize()])
    for nivel in NIVELES_ALERTA:
        filas.append(['niveles_alerta', nivel, nivel.capitalize()])
    filas.append(['si_no', 'si', 'Sí'])
    filas.append(['si_no', 'no', 'No'])

    _escribir_hoja(wb, 'choices', ['list_name', 'name', 'label'], filas)


def generar_xlsform_noticias(ruta_salida='xlsforms/noticias_emergencia_vzla.xlsx'):
    directorio = os.path.dirname(ruta_salida)
    if directorio:
        os.makedirs(directorio, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    survey = [
        ['date', 'fecha', 'Fecha de la noticia'],
        ['text', 'titulo', 'Título'],
        ['text', 'enlace', 'Enlace'],
        ['select_one estados', 'estado', 'Estado'],
        ['select_one tipos_evento', 'tipo_evento', 'Tipo de evento'],
        ['text', 'medio', 'Medio / fuente'],
        ['select_one niveles_alerta', 'nivel_alerta_asociado', 'Nivel de alerta asociado'],
        ['text', 'criterio_busqueda', 'Criterio de búsqueda (trazabilidad)'],
        ['today', 'fecha_carga', 'Fecha de carga a Kobo'],
    ]
    _escribir_hoja(wb, 'survey', ['type', 'name', 'label'], survey)
    _hoja_choices_comun(wb)
    _escribir_hoja(
        wb,
        'settings',
        ['form_title', 'form_id', 'version'],
        [['noticias_emergencia_vzla', 'noticias_emergencia_vzla', '1.0']],
    )

    wb.save(ruta_salida)
    print(f'Generado: {ruta_salida}')


def generar_xlsform_metricas_alerta(ruta_salida='xlsforms/metricas_alerta_vzla.xlsx'):
    directorio = os.path.dirname(ruta_salida)
    if directorio:
        os.makedirs(directorio, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    survey = [
        ['dateTime', 'fecha_calculo', 'Fecha de cálculo'],
        ['select_one estados', 'estado', 'Estado'],
        ['select_one tipos_evento', 'tipo_evento', 'Tipo de evento'],
        ['integer', 'ventana_horas', 'Ventana de tiempo (horas)'],
        ['integer', 'n_noticias', 'Número de noticias'],
        ['integer', 'n_fuentes_distintas', 'Fuentes distintas'],
        ['select_one si_no', 'tiene_victimas', '¿Hay víctimas mencionadas?'],
        ['select_one si_no', 'tiene_rescate_activo', '¿Hay rescate activo?'],
        ['select_one si_no', 'tiene_danios_estructurales', '¿Hay daños estructurales?'],
        ['dateTime', 'fecha_mas_reciente', 'Fecha de la noticia más reciente'],
        ['select_one niveles_alerta', 'nivel_alerta', 'Nivel de alerta'],
        ['text', 'urls_relacionadas', 'URLs relacionadas'],
    ]
    _escribir_hoja(wb, 'survey', ['type', 'name', 'label'], survey)
    _hoja_choices_comun(wb)
    _escribir_hoja(
        wb,
        'settings',
        ['form_title', 'form_id', 'version'],
        [['metricas_alerta_vzla', 'metricas_alerta_vzla', '1.0']],
    )

    wb.save(ruta_salida)
    print(f'Generado: {ruta_salida}')


def validar_xlsform(ruta):
    wb = openpyxl.load_workbook(ruta, data_only=True)

    print("\nValidación XLSForm")
    print("------------------")

    for hoja in ['survey', 'choices', 'settings']:
        if hoja not in wb.sheetnames:
            print(f'ERROR: falta hoja {hoja}')
            continue

        ws = wb[hoja]

        print(f'\nHoja: {hoja}')
        print(f'Filas: {ws.max_row}')
        print(f'Columnas: {ws.max_column}')

        for fila in ws.iter_rows(values_only=True):
            print(fila)


if __name__ == '__main__':
    generar_xlsform_noticias()
    generar_xlsform_metricas_alerta()
    validar_xlsform('xlsforms/noticias_emergencia_vzla.xlsx')